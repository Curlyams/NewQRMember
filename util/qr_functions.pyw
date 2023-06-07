import pandas as pd
import openpyxl as op
import openpyxl.utils


def members_with_codes(ws,col ='D'):
    """
    Takes all the member ID in master QR sheet and seperates them to a list

    :param ws: the excel worksheet found in the master qr sheet.
    :type ws: Excel worksheet
    :param col: chooses the col where the member id is located in ws default = 'D' 
    :type col: str
    return: member_id_lst a list with member ids.
    rtype: str
    """
    member_id_lst = []
    for cell in ws[col]:
        member_id_lst.append(cell.value)
    member_id_lst = [value for value in member_id_lst if value is not None]
    member_id_lst = member_id_lst[1:]
    return member_id_lst


def find_last_row(ws, min_row_=4,min_col_=4,max_col_=4):
    """
    Finds what row is empty in excel worksheet.

    :param ws: excel work sheet
    :tpye ws: excel work sheet
    :params: min_row_, min_col_, max_col_ : coordinates for where to start counting rows default is 4
    :typs coords: int
    :returns: row_num + min_row_ which is the index of last empty row
    :rtype: int
    """
    row_num = 0
    for col in ws.iter_cols(min_row=min_row_,min_col=min_col_,max_col=max_col_):
        for cell in col:
            if cell.value is not None:
                row_num += 1
            else:
                return row_num + min_row_
            
def add_members_data_to_sheet(new_member_df, ws):
    row_num = find_last_row(ws)

    for index, row in new_member_df.iterrows():
        ws[openpyxl.utils.get_column_letter(1) + str(row_num)] = row['Action Needed']
        ws[openpyxl.utils.get_column_letter(2) + str(row_num)] = row['Client First Name']
        ws[openpyxl.utils.get_column_letter(3) + str(row_num)] = row['Client Last Name']
        ws[openpyxl.utils.get_column_letter(4) + str(row_num)] = row['Customer Number']
        ws[openpyxl.utils.get_column_letter(5) + str(row_num)] = row['Language']
        row_num += 1


        
def load_export_df(export_path):
    export_df = pd.read_csv(export_path, low_memory=False).astype(str)
    #export_df['Drop-off Address']= export_df['Drop-off Street Number'].apply(str)+ ' ' + export_df['Drop-off Street'].apply(str)
    #export_df['Pick-up Address']= export_df['Pick-up Street Number'].apply(str)+ ' ' + export_df['Pick-up Street'].apply(str)
    return export_df

def load_master_qr_sheet(master_qr_path):
    member_wb = op.load_workbook(master_qr_path)
    all_member_ws = member_wb.active
    return all_member_ws,member_wb

def all_members_at_clinics(export_df,clinic_address):

    mr_members = export_df.query("`Provider` == 'Reimbursement Mileage'")
    #member_at_participating_clinics = mr_members.query("`Drop-off Address` in @clinic_address or `Pick-up Address` in @clinic_address and `Provider` == 'Reimbursement Mileage'")
    member_at_participating_clinics = mr_members.query("`Drop-off Street Number` in @clinic_address or `Pick-up Street Number` in @clinic_address")
    member_at_participating_clinics = member_at_participating_clinics.drop_duplicates(subset=['Customer Number'], keep ='first')
    return member_at_participating_clinics


def new_member_finder(all_member_ws,member_at_participating_clinics):
        member_id_lst = members_with_codes(all_member_ws)
    

    # Apply a filter to the DataFrame
        new_members = member_at_participating_clinics[~member_at_participating_clinics['Customer Number'].isin(member_id_lst)]

        # Check if the filtered DataFrame is empty
        if new_members.empty:
            new_members = pd.DataFrame()
            return new_members
        else:
            new_members.loc[:, 'Action Needed'] = 'Print QR Code Card'
            new_members = new_members.reindex(columns=['Action Needed','Client First Name','Client Last Name','Customer Number','Language'])
            return new_members
    